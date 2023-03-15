#! python3
# excelTextFileToSpreadsheet.py - Reads the contents of several text files and inserts those contents into a spreadsheet, with one line of text per row. The lines of the first text file will be in the cells of column A, the lines of the second text file will be in the cells of column B, and so on.

# Line 12, 13, and 29 - Folder and file names need to be customized in order to work for you.

import sys, os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#open text files
fileone = open("C:\\Users\OWNER\mu_code\madliblayout.txt", mode="rt")
filetwo = open("C:\\Users\OWNER\mu_code\myTestFolder\spam001.txt", mode="rt")

#read text files and returns list of strings, one strng per line of text
onestrings = fileone.readlines()
twostrings = filetwo.readlines()

#create workbook
wb = Workbook()
ws = wb.active #get default sheet. Always starts with one sheet.

#put strings into cells
for index, lines in enumerate(onestrings):
    ws['A' + str(index + 1)] = onestrings[index]
for index, lines in enumerate(twostrings):
    ws['B' + str(index + 1)] = twostrings[index]

wb.save('C:\\Users\\OWNER\\mu_code\\myTestFolder\\texttosheet.xlsx')
