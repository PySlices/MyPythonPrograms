#! python3
# excelMultiplicationTableMaker.py - makes a multiplication table

# Line 44 and 45 - Folder and file names need to be customized in order to work for you.
# In CMD, type 'py excelMultiplicationTableMaker.py X' (x= number to create a multiplication table from x * x)


import sys
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

#cmd line arguments
if len(sys.argv) < 1:
    print('Usage: py excelMultiplicationTableMaker.py [number]')
    sys.exit()
number = sys.argv[1]

#create excel doc
wb = Workbook() #create workbook
ws = wb.active #get default sheet. Always starts with one sheet.
ws.title = 'Multiplication Table' #name sheet


#create multiplication table headers
BoldFont = Font(bold=True) # Create a font.
rownumbers = 1
for i in range(2, int(number) + 2):
    ws['A' + str(i)] = rownumbers
    rownumbers += 1
    ws['A' + str(i)].font = BoldFont
columnnumbers = 1
for i in range(2, int(number) + 2):
    ws[str(get_column_letter(int(i))) + '1'] = columnnumbers
    columnnumbers += 1
    ws[str(get_column_letter(int(i))) + '1'].font = BoldFont

#generate multiplaction table content
for row in range(2, int(number) + 2): #row
    for column in range(2, int(number) + 2): #column
        ws[str(get_column_letter(int(column))) + str(row)] = (int(column) - 1) * (int(row) - 1)


wb.save('C:\\Users\OWNER\mu_code\myTestFolder\multiplicationTable.xlsx')
print('multiplicationTable.xlsx was created in ' + 'C:\\Users\OWNER\mu_code\myTestFolder')
