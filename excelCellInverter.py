#! python3
# excelCellInverter.py - instead of a vertical list, it becomes horizontal. For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa)
# Line 14, 41, and 42 - file and folder names need to be customized to work for you

# Uses list of lists [sheetData, x, y] for the cell at column x and row y


import sys, os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

#Load workbook. Enter the file path for the excel file to modify.
filename = 'C:\\Users\\OWNER\\mu_code\\myTestFolder\\automate_online-materials\\duesrecords.xlsx'
wb = load_workbook(filename)
ws = wb.active #get default sheet. Always starts with one sheet.

#write data to a list of lists
WBData = [] #sheetData[x][y]
for row in range(1, ws.max_row + 1):
    for column in range(1, ws.max_column + 1):
        sheetdata = ws[str(get_column_letter(column)) + str(row)].value
        x = column
        y = row
        eachlist = []
        eachlist.append(sheetdata)
        eachlist.append(x)
        eachlist.append(y)
        WBData.append(eachlist)

#Rewrite data inverted
index = 0
for i in WBData:
    ws[get_column_letter(WBData[index][2]) + str(WBData[index][1])] = WBData[index][0]
    index += 1

#save workbook.
print('Saving...')
basename = os.path.basename(filename)[0:-5]
#Enter the folder path where you want it saved on line 41 and 42.
wb.save('C:\\Users\\OWNER\\mu_code\\myTestFolder\\' + basename + '_invertedcopy.xlsx')
print(basename + '_copy.xlsx was created in ' + 'C:\\Users\OWNER\mu_code\myTestFolder')
